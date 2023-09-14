import openpyxl
import os

from src.models.info import Info
from src.models.assignment_group import AssignmentGroup
from src.models.assignment import Assignment
from src.models.page import Page
from src.models.module import Module
from src.models.quiz import Quiz
from src.models.module_content import ModuleContent


def if_empty_throw_error(value, error_message):
    """ Throws an error if the value is empty """
    if value is None or value == "":
        raise Exception(error_message)


def read_file_content(file_path):
    """ Reads the content of a file and returns it as a string """
    with open(file_path, 'r') as file:
        content = file.read()
    file.close()
    return content


def if_file_not_found_throw_error(file_path, error_message):
    """ Throws an error if the file is not found """
    if not os.path.isfile(file_path):
        raise Exception(error_message + f"\nPath to file: {file_path}")


class ConfigReader:
    def __init__(self, path_to_config_folder):
        self.path_to_config_folder = path_to_config_folder
        self.path_to_xlsx = os.path.join(path_to_config_folder, "config.xlsx")
        pass

    @staticmethod
    def check_info_sheet(self, workbook):
        info_sheet = workbook["Algemene info"]

        if_empty_throw_error(info_sheet["B1"].value, "Sheet 'Algemene info', cell B1: course ID is empty")
        if_empty_throw_error(info_sheet["B2"].value, "Sheet 'Algemene info', cell B2: course image is empty")
        if_empty_throw_error(info_sheet["B3"].value, "Sheet 'Algemene info', cell B3: teachers are empty")
        if_empty_throw_error(info_sheet["B4"].value, "Sheet 'Algemene info', cell B3: contacts are empty")
        if_empty_throw_error(info_sheet["B5"].value, "Sheet 'Algemene info', cell B5: course name is empty")

        return Info(
            course_id=info_sheet["B1"].value,
            course_image=os.path.join(self.path_to_config_folder, 'afbeeldingen', info_sheet["B2"].value),
            course_name=info_sheet["B5"].value,
            teachers=info_sheet["B3"].value,
            contacts=info_sheet["B4"].value,
        )

    @staticmethod
    def check_assignment_groups_sheet(self, workbook):
        assignment_groups = []
        assignment_groups_sheet = workbook["Opdrachtgroepen"]
        for row in assignment_groups_sheet.iter_rows(min_row=2, values_only=True):
            # If name is empty, skip this row
            if row[1] == None or row[1] == "":
                continue

            assignment_groups.append(
                AssignmentGroup(
                    name=row[0],
                    weight=row[1]
                )
            )
        return assignment_groups

    @staticmethod
    def check_assignments_sheet(self, workbook):
        assignments = []
        assignments_sheet = workbook["Opdrachten"]
        for row in assignments_sheet.iter_rows(min_row=2, values_only=True):
            # If name is empty, skip this row
            if row[1] == None or row[1] == "":
                continue

            assignment_description_file = row[7]
            # Check if the assignment description file exists
            if assignment_description_file != None and assignment_description_file != "":
                if_file_not_found_throw_error(
                    os.path.join(self.path_to_config_folder, 'opdrachten',
                                 assignment_description_file),
                    f"Assignment description file {assignment_description_file} not found")

            assignments.append(
                Assignment(
                    assignment_group=row[0],
                    name=row[1],
                    due_at=row[2],
                    available_at=row[3],
                    available_until=row[4],
                    points=row[5],
                    published=row[6],
                    description_content='' if assignment_description_file is None else read_file_content(
                        os.path.join(self.path_to_config_folder, 'opdrachten', assignment_description_file))
                ))

        return assignments

    @staticmethod
    def check_quizzes_sheet(self, workbook):
        quizzes = []
        quizzes_sheet = workbook["Toetsen"]
        for row in quizzes_sheet.iter_rows(min_row=2, values_only=True):
            # If name is empty, skip this row
            if row[1] == None or row[1] == "":
                continue

            quizzes.append(
                Quiz(
                    assignment_group=row[0],
                    title=row[1],
                    due_at=row[2],
                    available_at=row[3],
                    available_until=row[4],
                    points=row[5],
                    published=row[6],
                ))
        return quizzes

    @staticmethod
    def check_pages_sheet(self, workbook):
        pages = []
        pages_sheet = workbook["Pagina's"]
        for row in pages_sheet.iter_rows(min_row=2, values_only=True):
            # If the id is empty, skip this row
            if row[1] == None or row[1] == "":
                continue

            page_content_file = row[2]
            # Check if the assignment description file exists
            if page_content_file != None and page_content_file != "":
                if not os.path.isfile(os.path.join(self.path_to_config_folder, 'paginas', page_content_file)):
                    raise Exception(
                        f"Page content file {page_content_file} not found")
            pages.append(Page(
                title=row[0],
                published=row[1],
                body='' if page_content_file == None else read_file_content(
                    os.path.join(self.path_to_config_folder, 'paginas', page_content_file)),
            ))
        return pages

    @staticmethod
    def check_modules_sheet(self, workbook):
        modules = []
        modules_sheet = workbook["Modules"]
        for row in modules_sheet.iter_rows(min_row=2, values_only=True):
            # If name is empty, skip this row
            if row[1] == None or row[1] == "":
                continue

            modules.append(
                Module(
                    name=row[0],
                    published=row[1]
                ))
        return modules

    @staticmethod
    def check_module_contents_sheet(self, workbook, modules):
        module_contents_sheet = workbook["Module inhoud"]
        for row in module_contents_sheet.iter_rows(min_row=2, values_only=True):
            # If the module is empty, skip this row
            if row[0] == None or row[0] == "":
                continue
            # Find the module with the name
            module = next((m for m in modules if m.name == row[0]), None)
            if module is None:
                raise Exception(f"Module with name {row[0]} not found")

            file_path = row[7]
            # If it is a file, check if the file exists
            if file_path is not None and file_path != "":
                if_file_not_found_throw_error(
                    os.path.join(self.path_to_config_folder, 'bestanden', file_path),
                    f"File {file_path} not found")
                file_path = os.path.join(self.path_to_config_folder, 'bestanden', file_path)

            # Add the content to the module
            module.contents.append(
                ModuleContent(
                    type=row[1],
                    display_name=row[2],
                    published=row[3],
                    assignment=row[4],
                    url=row[5],
                    quiz=row[6],
                    file_path=file_path,
                    indent=row[8] if row[8] is not None else 0,
                ))

        return modules

    def check_config(self):
        workbook = openpyxl.load_workbook(self.path_to_xlsx)
        info = self.check_info_sheet(self, workbook)
        assignment_groups = self.check_assignment_groups_sheet(self, workbook)
        assignments = self.check_assignments_sheet(self, workbook)
        quizzes = self.check_quizzes_sheet(self, workbook)
        pages = self.check_pages_sheet(self, workbook)
        modules = self.check_modules_sheet(self, workbook)
        modules = self.check_module_contents_sheet(self, workbook, modules)

        return info, assignment_groups, assignments, quizzes, modules, pages
